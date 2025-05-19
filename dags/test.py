import configparser
from airflow import DAG
from airflow.decorators import task, dag
from airflow.operators.empty import EmptyOperator
from datetime import datetime, timedelta
import pandas as pd
import oracledb
import smtplib
from pythainlp.util import thai_strftime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Color, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pendulum
import logging

# ตกแต่ง Excel
red_font = Font(color="FF0000")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_worksheet(ws):
    
    # จัดรูปแบบหัวข้อรายงาน (แถวที่ 1)
    # ws.merge_cells('A1:H1')  # รวมเซลล์ A1 ถึง H1
    # title_cell = ws['A1']
    # title_cell.value = 'รายงานประจำวัน'
    # title_cell.font = Font(bold=True, size=14)
    # title_cell.alignment = Alignment(horizontal='center', vertical='center')
    # title_cell.border = thin_border
      
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = yellow_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # ข้อมูล (row 2 เป็นต้นไป)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value in [None, "", "NaN", float("nan")]:
                cell.value = "-"
                cell.font = red_font
            else:
                cell.font = Font(color="000000")  # สีดำปกติ
            cell.alignment = Alignment(horizontal='left')
            cell.border = thin_border

    # ปรับความกว้างคอลัมน์
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

# อ่าน config
config_file_path = 'config/db_config.cfg'
config = configparser.ConfigParser()
config.read(config_file_path)
local = config.get("variable", "tz")
local_tz = pendulum.timezone(local)
currentDateAndTime = datetime.now(tz=local_tz)
currentDate = currentDateAndTime.strftime("%Y-%m-%d")
currentDateFormat = currentDateAndTime.strftime("%d-%m-%y")
currentthaiDate = thai_strftime(currentDateAndTime, "%Aที่ %-d %B พ.ศ.%Y")

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2023, 1, 1, tzinfo=local_tz),
    'retries': 3,
    'retry_delay': timedelta(minutes=5),
    'email_on_failure': False,
    'email_on_retry': False,
}

@dag(
    schedule_interval='0 20 * * *',
    start_date=datetime(2023, 1, 1, tzinfo=local_tz),
    catchup=False,
    default_args=default_args,
    tags=['report'],
    max_active_runs=1,
)
def opportunity_report_dag():
    
    @task(retries=20, retry_delay=timedelta(minutes=2))
    def query_data() -> pd.DataFrame:
        """ดึงข้อมูลจาก Oracle Database"""
        try:
            env = 'dev'
            db_host = config.get(env, 'host')
            db_port = config.getint(env, 'port')
            db_username = config.get(env, 'username')
            db_password = config.get(env, 'password')
            db_name = config.get(env, 'dbname')

            logging.info(f"Connecting to Oracle DB: {db_host}:{db_port}/{db_name}")
            
            dsn_name = oracledb.makedsn(db_host, db_port, service_name=db_name)
            with oracledb.connect(
                user=db_username,
                password=db_password,
                dsn=dsn_name,
            ) as conn:
                cursor = conn.cursor()
                query = """
                    SELECT * FROM SYSBYTEDES s 
                    WHERE TABLENAME LIKE 'PRODUCTCORP'
                """
                cursor.execute(query)
                
                columns = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df = pd.DataFrame(data, columns=columns)
                
                logging.info(f"Retrieved {len(df)} records successfully")
                return df
                
        except Exception as e:
            logging.error(f"Error querying data: {str(e)}")
            raise
        
    @task(retries=20, retry_delay=timedelta(minutes=2))
    def query_data_sheet2() -> pd.DataFrame:
        """ดึงข้อมูลจาก Oracle Database"""
        try:
            env = 'dev'
            db_host = config.get(env, 'host')
            db_port = config.getint(env, 'port')
            db_username = config.get(env, 'username')
            db_password = config.get(env, 'password')
            db_name = config.get(env, 'dbname')

            logging.info(f"Connecting to Oracle DB: {db_host}:{db_port}/{db_name}")
            
            dsn_name = oracledb.makedsn(db_host, db_port, service_name=db_name)
            with oracledb.connect(
                user=db_username,
                password=db_password,
                dsn=dsn_name,
            ) as conn:
                cursor = conn.cursor()
                query = """
                    SELECT * FROM SYSBYTEDES s 
                    WHERE TABLENAME LIKE 'LEADSEARCH'
                """
                cursor.execute(query)
                
                columns = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df = pd.DataFrame(data, columns=columns)
                
                logging.info(f"Retrieved {len(df)} records successfully")
                return df
                
        except Exception as e:
            logging.error(f"Error querying data: {str(e)}")
            raise

    @task
    def generate_excel_report(df_product: pd.DataFrame, df_lead: pd.DataFrame) -> str:
        """สร้างรายงาน Excel จากข้อมูล"""
        try:
            # สร้าง workbook และ worksheet
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "ProductList"
            
            # เพิ่มข้อมูลลงใน worksheet 1
            for r in dataframe_to_rows(df_product, index=False, header=True):
                ws1.append(r)
            
            # สร้าง worksheet 2
            ws2 = wb.create_sheet(title="LeadSearch")
            
            # เพิ่มข้อมูลลงใน worksheet 2
            for r in dataframe_to_rows(df_lead, index=False, header=True):
                ws2.append(r)
            
            # หลังจากเพิ่มข้อมูลใน ws1
            format_worksheet(ws1)

            # หลังจากเพิ่มข้อมูลใน ws2
            format_worksheet(ws2)

            # บันทึกไฟล์
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            excel_report = base64.b64encode(excel_file.read()).decode('utf-8')
            
            logging.info("Excel report generated successfully")
            return excel_report
            
        except Exception as e:
            logging.error(f"Error generating report: {str(e)}")
            raise

    @task(
        retries=3,
        retry_delay=timedelta(minutes=2),
        execution_timeout=timedelta(minutes=10)
    )
    def send_email(excel_report: str) -> bool:
        """ส่งอีเมลพร้อมไฟล์แนบ"""
        try:
            smtp_host = config.get('variable', 'smtp_host')
            smtp_port = config.getint('variable', 'smtp_port')
            smtp_mail = config.get('variable', 'sender_email')
            smtp_password = config.get('variable', 'sender_email_password')
            receiver_emails = config.get('variable', 'receiver_emails').split(", ")
            cc_emails = config.get('variable', 'cc_emails').split(", ")
            
            logging.info(f"Preparing to send email via {smtp_host}:{smtp_port}")
            
            msg = MIMEMultipart()
            msg['From'] = smtp_mail
            msg['To'] = ", ".join(receiver_emails)
            msg['Cc'] = ", ".join(cc_emails)
            msg['Subject'] = f"รายงาน ProductList และ LeadSearch ประจำวันที่ {currentthaiDate}"
            
            # เนื้อหาอีเมล
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif;">
                    <p>เรียน ทีมงาน</p>
                    <p>ส่งรายงาน ProductList และ LeadSearch ประจำวันที่ <strong>{currentthaiDate}</strong></p>
                    <p>รายงานสรุปข้อมูลตามไฟล์แนบ</p>
                    <p style="color: red;">อีเมลนี้ถูกส่งโดยระบบอัตโนมัติ</p>
                    <p>ขอแสดงความนับถือ<br>ทีม IT</p>
                </body>
            </html>
            """
            msg.attach(MIMEText(body, 'html'))
            
            # แนบไฟล์ Excel
            excel_data = base64.b64decode(excel_report)
            attachment = MIMEApplication(
                excel_data,
                Name=f"Product_Lead_Report_{currentDateFormat}.xlsx"
            )
            attachment['Content-Disposition'] = f'attachment; filename="Product_Lead_Report_{currentDateFormat}.xlsx"'
            msg.attach(attachment)
            
            # ส่งอีเมล
            if smtp_port == 465:
                with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30) as server:
                    server.login(smtp_mail, smtp_password)
                    server.send_message(msg)
            else:
                with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
                    server.starttls()
                    server.login(smtp_mail, smtp_password)
                    server.send_message(msg)
            
            logging.info(f"Email sent successfully to {receiver_emails} (CC: {cc_emails})")
            return True
            
        except Exception as e:
            logging.error(f"Error sending email: {str(e)}")
            raise

    # กำหนด Flow
    start = EmptyOperator(task_id='start')
    end = EmptyOperator(task_id='end')
    
    data_product = query_data()
    data_lead = query_data_sheet2()
    report = generate_excel_report(data_product, data_lead)
    email = send_email(report)
    
    start >> [data_product, data_lead] >> report >> email >> end

# สร้าง DAG
opportunity_report = opportunity_report_dag()